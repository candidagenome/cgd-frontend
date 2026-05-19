#!/usr/bin/env python3
"""
Extract Alignment Statistics from HISAT2 Logs
==============================================
Parses HISAT2 alignment logs and updates metadata (XLSX or TSV) with alignment percentages.

Usage:
    python extract_alignment_stats.py <metadata.xlsx> <logs_dir> [--output <output.xlsx>]

Example:
    python extract_alignment_stats.py Iracane_2021_metadata.xlsx /data/tmp/aligned/
    python extract_alignment_stats.py Iracane_2021_metadata.xlsx ./logs --output updated_metadata.xlsx

The script looks for log files named {SRR_ID}.log or {SRR_ID}_hisat2.log in the logs directory.
"""

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


ALIGNMENT_THRESHOLD = 85.0


def parse_hisat2_log(log_path: Path) -> dict:
    """Parse HISAT2 log file and extract alignment statistics."""
    stats = {
        'total_reads': None,
        'aligned_0': None,
        'aligned_1': None,
        'aligned_multi': None,
        'overall_rate': None,
    }

    if not log_path.exists():
        return stats

    content = log_path.read_text()

    # Total reads
    match = re.search(r'(\d+) reads; of these:', content)
    if match:
        stats['total_reads'] = int(match.group(1))

    # Aligned 0 times (unaligned)
    match = re.search(r'(\d+) \([\d.]+%\) aligned (?:concordantly )?0 times', content)
    if match:
        stats['aligned_0'] = int(match.group(1))

    # Aligned exactly 1 time
    match = re.search(r'(\d+) \([\d.]+%\) aligned (?:concordantly )?exactly 1 time', content)
    if match:
        stats['aligned_1'] = int(match.group(1))

    # Aligned >1 times
    match = re.search(r'(\d+) \([\d.]+%\) aligned (?:concordantly )?>1 times', content)
    if match:
        stats['aligned_multi'] = int(match.group(1))

    # Overall alignment rate
    match = re.search(r'([\d.]+)% overall alignment rate', content)
    if match:
        stats['overall_rate'] = float(match.group(1))

    return stats


def find_log_file(logs_dir: Path, srr_id: str) -> Path:
    """Find the log file for a given SRR ID."""
    patterns = [
        f"{srr_id}.log",
        f"{srr_id}_hisat2.log",
        f"{srr_id}_align.log",
        f"{srr_id}.hisat2.log",
    ]

    for pattern in patterns:
        log_path = logs_dir / pattern
        if log_path.exists():
            return log_path

    # Try looking in subdirectory
    subdir = logs_dir / srr_id
    if subdir.exists():
        for pattern in patterns:
            log_path = subdir / pattern
            if log_path.exists():
                return log_path

    return None


def update_metadata_xlsx(input_path: str, logs_dir: str, output_path: str = None) -> dict:
    """Update metadata XLSX with alignment statistics."""
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl is required for xlsx files. Install with: pip install openpyxl")
        sys.exit(1)

    from openpyxl import Workbook
    from copy import copy

    logs_dir = Path(logs_dir)

    if output_path is None:
        output_path = input_path

    # Load the workbook (not read_only so we can modify)
    wb = load_workbook(input_path)

    stats_summary = {
        'total_samples': 0,
        'logs_found': 0,
        'logs_missing': 0,
        'passed': 0,
        'failed': 0,
    }

    if 'Samples' not in wb.sheetnames:
        print("ERROR: No 'Samples' sheet found in xlsx file")
        return stats_summary

    ws = wb['Samples']

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]

    # Add Align_Pct and Status columns if not present
    align_col = None
    status_col = None

    if 'Align_Pct' in headers:
        align_col = headers.index('Align_Pct') + 1
    else:
        align_col = len(headers) + 1
        ws.cell(row=1, column=align_col, value='Align_Pct')

    if 'Status' in headers:
        status_col = headers.index('Status') + 1
    else:
        status_col = len(headers) + 2 if align_col == len(headers) + 1 else len(headers) + 1
        ws.cell(row=1, column=status_col, value='Status')

    # Find SRR_ID column
    srr_col = headers.index('SRR_ID') + 1 if 'SRR_ID' in headers else 1

    # Process each sample row
    for row_idx in range(2, ws.max_row + 1):
        srr_id = ws.cell(row=row_idx, column=srr_col).value
        if not srr_id:
            continue

        srr_id = str(srr_id).strip()
        stats_summary['total_samples'] += 1

        log_path = find_log_file(logs_dir, srr_id)

        if log_path:
            stats_summary['logs_found'] += 1
            log_stats = parse_hisat2_log(log_path)

            if log_stats['overall_rate'] is not None:
                align_pct = log_stats['overall_rate']
                ws.cell(row=row_idx, column=align_col, value=round(align_pct, 1))

                if align_pct >= ALIGNMENT_THRESHOLD:
                    ws.cell(row=row_idx, column=status_col, value='OK')
                    stats_summary['passed'] += 1
                else:
                    ws.cell(row=row_idx, column=status_col, value='FAILED')
                    stats_summary['failed'] += 1

                print(f"  {srr_id}: {align_pct:.1f}% {'OK' if align_pct >= ALIGNMENT_THRESHOLD else 'FAILED'}")
            else:
                print(f"  {srr_id}: Could not parse alignment rate from {log_path}")
                stats_summary['logs_missing'] += 1
        else:
            print(f"  {srr_id}: Log file not found")
            stats_summary['logs_missing'] += 1

    wb.save(output_path)
    wb.close()

    return stats_summary


def update_metadata_tsv(input_path: str, logs_dir: str, output_path: str = None) -> dict:
    """Update metadata TSV with alignment statistics."""
    logs_dir = Path(logs_dir)

    if output_path is None:
        output_path = input_path

    with open(input_path, 'r') as f:
        lines = f.readlines()

    output_lines = []
    in_sample_section = False
    headers = []

    stats_summary = {
        'total_samples': 0,
        'logs_found': 0,
        'logs_missing': 0,
        'passed': 0,
        'failed': 0,
    }

    for i, line in enumerate(lines):
        stripped = line.strip()

        if stripped.startswith('## SAMPLE METADATA'):
            in_sample_section = True
            output_lines.append(line)
            continue

        if stripped.startswith('## ') and in_sample_section:
            in_sample_section = False

        if in_sample_section and stripped and not stripped.startswith('#'):
            parts = line.rstrip('\n').split('\t')

            if not headers:
                headers = parts
                if 'Align_Pct' not in headers:
                    headers.append('Align_Pct')
                if 'Status' not in headers:
                    headers.append('Status')
                output_lines.append('\t'.join(headers) + '\n')
                continue

            srr_id = parts[0] if parts else ''

            if srr_id:
                stats_summary['total_samples'] += 1

                while len(parts) < len(headers):
                    parts.append('')

                align_idx = headers.index('Align_Pct') if 'Align_Pct' in headers else -1
                status_idx = headers.index('Status') if 'Status' in headers else -1

                log_path = find_log_file(logs_dir, srr_id)

                if log_path:
                    stats_summary['logs_found'] += 1
                    log_stats = parse_hisat2_log(log_path)

                    if log_stats['overall_rate'] is not None:
                        align_pct = log_stats['overall_rate']
                        parts[align_idx] = f"{align_pct:.1f}"

                        if align_pct >= ALIGNMENT_THRESHOLD:
                            parts[status_idx] = 'OK'
                            stats_summary['passed'] += 1
                        else:
                            parts[status_idx] = 'FAILED'
                            stats_summary['failed'] += 1

                        print(f"  {srr_id}: {align_pct:.1f}% {'OK' if align_pct >= ALIGNMENT_THRESHOLD else 'FAILED'}")
                    else:
                        print(f"  {srr_id}: Could not parse alignment rate from {log_path}")
                        stats_summary['logs_missing'] += 1
                else:
                    print(f"  {srr_id}: Log file not found")
                    stats_summary['logs_missing'] += 1

                output_lines.append('\t'.join(parts) + '\n')
            else:
                output_lines.append(line)
        else:
            output_lines.append(line)

    with open(output_path, 'w') as f:
        f.writelines(output_lines)

    return stats_summary


def update_metadata(input_path: str, logs_dir: str, output_path: str = None) -> dict:
    """Update metadata file (XLSX or TSV) with alignment statistics."""
    input_lower = input_path.lower()

    if input_lower.endswith('.xlsx'):
        return update_metadata_xlsx(input_path, logs_dir, output_path)
    else:
        return update_metadata_tsv(input_path, logs_dir, output_path)


def main():
    parser = argparse.ArgumentParser(
        description='Extract alignment stats from HISAT2 logs and update metadata file',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python extract_alignment_stats.py metadata.xlsx /data/tmp/aligned/
  python extract_alignment_stats.py metadata.xlsx ./logs --output updated.xlsx
        '''
    )
    parser.add_argument('metadata_file', help='Path to metadata file (XLSX or TSV)')
    parser.add_argument('logs_dir', help='Directory containing HISAT2 log files')
    parser.add_argument('--output', '-o', help='Output file path (default: update in place)')
    parser.add_argument('--threshold', '-t', type=float, default=85.0,
                        help='Alignment threshold percentage (default: 85.0)')

    args = parser.parse_args()

    global ALIGNMENT_THRESHOLD
    ALIGNMENT_THRESHOLD = args.threshold

    print(f"Extracting alignment stats from: {args.logs_dir}")
    print(f"Alignment threshold: {ALIGNMENT_THRESHOLD}%")
    print()

    summary = update_metadata(args.metadata_file, args.logs_dir, args.output)

    output_file = args.output or args.metadata_file
    print()
    print(f"Updated: {output_file}")
    print()
    print("=== Summary ===")
    print(f"Total samples:  {summary['total_samples']}")
    print(f"Logs found:     {summary['logs_found']}")
    print(f"Logs missing:   {summary['logs_missing']}")
    print(f"Passed (>={ALIGNMENT_THRESHOLD}%): {summary['passed']}")
    print(f"Failed (<{ALIGNMENT_THRESHOLD}%): {summary['failed']}")


if __name__ == '__main__':
    main()

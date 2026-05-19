#!/usr/bin/env python3
"""
RNA-seq Dataset Import Script
=============================
Reads curator-provided metadata (XLSX or TSV) and generates:
1. JBrowse2 track configuration (JSON)
2. Expression service configuration (Python snippet)
3. JBrowse2 symlink creation script (Bash)

Usage:
    python import_rnaseq.py <metadata.xlsx> [--output-dir <dir>]

Example:
    python import_rnaseq.py Shivarathri_2022_metadata.xlsx --output-dir ./output
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


ALIGNMENT_THRESHOLD = 85.0  # Minimum alignment percentage


def parse_metadata_xlsx(filepath: str) -> tuple[dict, list[dict]]:
    """Parse the metadata XLSX file.

    Expects two sheets:
    - "Study": Study metadata (Study_ID, PMID, Organism, Category)
    - "Samples": Sample metadata (SRR_ID, Condition_Label, Bucket, etc.)

    Returns:
        Tuple of (study_info dict, list of sample dicts)
    """
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl is required for xlsx files. Install with: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(filepath, read_only=True)

    # Parse Study sheet
    study_info = {}
    if 'Study' in wb.sheetnames:
        ws = wb['Study']
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            headers = [str(h).strip() if h else '' for h in rows[0]]
            values = rows[1]
            for i, header in enumerate(headers):
                if header and i < len(values) and values[i]:
                    study_info[header] = str(values[i]).strip()

    # Parse Samples sheet
    samples = []
    if 'Samples' in wb.sheetnames:
        ws = wb['Samples']
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(h).strip() if h else '' for h in rows[0]]
            for row in rows[1:]:
                sample = {}
                for i, header in enumerate(headers):
                    if header and i < len(row) and row[i] is not None:
                        sample[header] = str(row[i]).strip()
                if sample.get('SRR_ID'):
                    samples.append(sample)

    wb.close()
    return study_info, samples


def parse_metadata_tsv(filepath: str) -> tuple[dict, list[dict]]:
    """Parse the metadata TSV file.

    Returns:
        Tuple of (study_info dict, list of sample dicts)
    """
    study_info = {}
    samples = []
    current_section = None
    headers = []

    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()

            # Skip empty lines
            if not line:
                continue

            # Detect section headers
            if line.startswith('## STUDY METADATA'):
                current_section = 'study'
                continue
            elif line.startswith('## SAMPLE METADATA'):
                current_section = 'sample'
                continue

            # Skip comments
            if line.startswith('#'):
                continue

            # Parse tab-separated values
            parts = line.split('\t')

            if current_section == 'study':
                if not headers:
                    headers = parts
                else:
                    for i, val in enumerate(parts):
                        if i < len(headers) and val:
                            study_info[headers[i]] = val
                    headers = []  # Reset for sample section

            elif current_section == 'sample':
                if not headers:
                    headers = parts
                else:
                    sample = {}
                    for i, val in enumerate(parts):
                        if i < len(headers):
                            sample[headers[i]] = val
                    if sample.get('SRR_ID'):
                        samples.append(sample)

    return study_info, samples


def parse_metadata(filepath: str) -> tuple[dict, list[dict], list[dict]]:
    """Parse metadata file (XLSX or TSV).

    Returns:
        Tuple of (study_info dict, list of passed sample dicts, list of failed sample dicts)
    """
    filepath_lower = filepath.lower()

    if filepath_lower.endswith('.xlsx'):
        study_info, samples = parse_metadata_xlsx(filepath)
    elif filepath_lower.endswith('.tsv') or filepath_lower.endswith('.txt'):
        study_info, samples = parse_metadata_tsv(filepath)
    else:
        # Try to detect format
        try:
            study_info, samples = parse_metadata_xlsx(filepath)
        except Exception:
            study_info, samples = parse_metadata_tsv(filepath)

    # Separate passed and failed samples
    passed_samples = []
    failed_samples = []

    for sample in samples:
        status = sample.get('Status', '').upper()
        align_pct_str = sample.get('Align_Pct', '')

        # Auto-detect failure if Align_Pct is provided and below threshold
        if align_pct_str:
            try:
                align_pct = float(align_pct_str)
                if align_pct < ALIGNMENT_THRESHOLD:
                    status = 'FAILED'
                    sample['Status'] = 'FAILED'
                    sample['_fail_reason'] = f'Alignment {align_pct:.1f}% < {ALIGNMENT_THRESHOLD}% threshold'
            except ValueError:
                pass

        if status == 'FAILED':
            if '_fail_reason' not in sample:
                sample['_fail_reason'] = sample.get('Notes', 'Low quality')
            failed_samples.append(sample)
        else:
            passed_samples.append(sample)

    return study_info, passed_samples, failed_samples


def generate_jbrowse2_tracks(study_info: dict, samples: list[dict]) -> list[dict]:
    """Generate JBrowse2 track configuration (Coverage tracks only)."""
    tracks = []
    study_id = study_info.get('Study_ID', 'Unknown')
    prefix = study_id.replace('_', '')

    for sample in samples:
        srr_id = sample.get('SRR_ID', '')
        label = sample.get('Condition_Label', srr_id)

        # Coverage track only (BigWig) - no BAM/alignments tracks
        tracks.append({
            "type": "QuantitativeTrack",
            "trackId": f"{prefix}_{srr_id}_coverage",
            "name": f"{label} ({srr_id})",
            "adapter": {
                "type": "BigWigAdapter",
                "bigWigLocation": {
                    "uri": f"{prefix}_{srr_id}_coverage.bigwig",
                    "locationType": "UriLocation"
                }
            },
            "assemblyNames": [study_info.get('Organism', 'C_auris_B8441')],
            "category": ["RNA-seq", study_id],
            "displays": [{
                "type": "LinearWiggleDisplay",
                "displayId": f"{prefix}_{srr_id}_coverage_display"
            }]
        })

    return tracks


def generate_expression_config(study_info: dict, samples: list[dict]) -> str:
    """Generate Python snippet for expression_service.py."""
    study_id = study_info.get('Study_ID', 'Unknown')
    pmid = study_info.get('PMID', 'None')
    category = study_info.get('Category', 'Unknown')
    organism = study_info.get('Organism', 'C_auris_B8441')

    # Find control sample (first one with bucket=control)
    control_id = None
    for sample in samples:
        if sample.get('Bucket', '').lower() == 'control':
            control_id = sample.get('SRR_ID')
            break

    if not control_id and samples:
        control_id = samples[0].get('SRR_ID')

    # Build conditions dict
    conditions = []
    for sample in samples:
        srr_id = sample.get('SRR_ID', '')
        label = sample.get('Condition_Label', srr_id)
        bucket = sample.get('Bucket', 'basic_biology').lower()
        conditions.append(f'            "{srr_id}": {{"label": "{label}", "bucket": "{bucket}"}},')

    config = f'''
        "{study_id}": {{
            "category": "{category}",
            "pmid": "{pmid if pmid != 'None' else 'None'}",
            "path_style": "direct",
            "control": "{control_id}",
            "conditions": {{
{chr(10).join(conditions)}
            }},
        }},
'''

    return f"# Add to EXPRESSION_STUDIES['{organism}']:\n{config}"


def generate_symlink_script(study_info: dict, samples: list[dict]) -> str:
    """Generate Bash script for creating JBrowse2 symlinks."""
    study_id = study_info.get('Study_ID', 'Unknown')
    prefix = study_id.replace('_', '')
    organism = study_info.get('Organism', 'C_auris_B8441')

    script = f'''#!/bin/bash
# JBrowse2 Symlink Script for {study_id}
# Run on cgd-frontend-dev

set -e

JBROWSE_DIR=/data/jbrowse2
HTS_DIR=/data/HTS/{organism}/bam/{study_id}

# Backup config
cp $JBROWSE_DIR/config.json $JBROWSE_DIR/config.json.bak.$(date +%Y%m%d_%H%M%S)

echo "Creating symlinks for {study_id}..."

'''

    for sample in samples:
        srr_id = sample.get('SRR_ID', '')
        script += f'''# {sample.get('Condition_Label', srr_id)}
ln -sf $HTS_DIR/{srr_id}/{srr_id}_sorted_hits.bigwig $JBROWSE_DIR/{prefix}_{srr_id}_coverage.bigwig

'''

    script += f'''echo "Symlinks created."
echo ""
echo "Next: Add tracks from {study_id}_jbrowse2_tracks.json to config.json:"
echo "  jq '.tracks += input' $JBROWSE_DIR/config.json {study_id}_jbrowse2_tracks.json > config.json.new"
echo "  mv config.json.new $JBROWSE_DIR/config.json"
'''

    return script


def generate_failed_report(study_info: dict, failed_samples: list[dict]) -> str:
    """Generate a report of failed samples."""
    study_id = study_info.get('Study_ID', 'Unknown')

    report = f"""# Failed Samples Report - {study_id}
# Generated by import_rnaseq.py
# Samples below {ALIGNMENT_THRESHOLD}% alignment threshold

Total failed: {len(failed_samples)}

"""
    for sample in failed_samples:
        srr_id = sample.get('SRR_ID', '')
        label = sample.get('Condition_Label', srr_id)
        align_pct = sample.get('Align_Pct', 'N/A')
        reason = sample.get('_fail_reason', 'Unknown')
        report += f"- {srr_id}: {label}\n"
        report += f"  Alignment: {align_pct}%\n"
        report += f"  Reason: {reason}\n\n"

    return report


def main():
    parser = argparse.ArgumentParser(
        description='Import RNA-seq dataset from metadata file (XLSX or TSV)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python import_rnaseq.py Shivarathri_2022_metadata.xlsx
  python import_rnaseq.py Wang_2024_metadata.xlsx --output-dir ./output
        '''
    )
    parser.add_argument('metadata_file', help='Path to metadata file (XLSX or TSV)')
    parser.add_argument('--output-dir', '-o', default='.', help='Output directory (default: current)')

    args = parser.parse_args()

    # Parse metadata
    print(f"Reading metadata from: {args.metadata_file}")
    study_info, passed_samples, failed_samples = parse_metadata(args.metadata_file)

    if not study_info.get('Study_ID'):
        print("ERROR: No Study_ID found in metadata file")
        sys.exit(1)

    total_samples = len(passed_samples) + len(failed_samples)
    if total_samples == 0:
        print("ERROR: No samples found in metadata file")
        sys.exit(1)

    study_id = study_info['Study_ID']
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Study: {study_id}")
    print(f"Organism: {study_info.get('Organism', 'Unknown')}")
    print(f"PMID: {study_info.get('PMID', 'N/A')}")
    print(f"Samples: {len(passed_samples)} passed, {len(failed_samples)} failed (of {total_samples} total)")

    # Show failed samples
    if failed_samples:
        print()
        print(f"⚠ FAILED SAMPLES (alignment < {ALIGNMENT_THRESHOLD}%):")
        for sample in failed_samples:
            srr_id = sample.get('SRR_ID', '')
            align_pct = sample.get('Align_Pct', 'N/A')
            print(f"  - {srr_id}: {align_pct}% - {sample.get('_fail_reason', 'Failed')}")
    print()

    if not passed_samples:
        print("ERROR: No samples passed quality threshold")
        sys.exit(1)

    # Generate JBrowse2 tracks (only passed samples)
    tracks = generate_jbrowse2_tracks(study_info, passed_samples)
    tracks_file = output_dir / f"{study_id}_jbrowse2_tracks.json"
    with open(tracks_file, 'w') as f:
        json.dump(tracks, f, indent=2)
    print(f"Generated: {tracks_file} ({len(tracks)} tracks)")

    # Generate expression config (only passed samples)
    expr_config = generate_expression_config(study_info, passed_samples)
    expr_file = output_dir / f"{study_id}_expression_config.py"
    with open(expr_file, 'w') as f:
        f.write(expr_config)
    print(f"Generated: {expr_file}")

    # Generate symlink script (only passed samples)
    symlink_script = generate_symlink_script(study_info, passed_samples)
    script_file = output_dir / f"{study_id}_create_symlinks.sh"
    with open(script_file, 'w') as f:
        f.write(symlink_script)
    script_file.chmod(0o755)
    print(f"Generated: {script_file}")

    # Generate failed samples report if any
    if failed_samples:
        failed_report = generate_failed_report(study_info, failed_samples)
        failed_file = output_dir / f"{study_id}_FAILED_SAMPLES.txt"
        with open(failed_file, 'w') as f:
            f.write(failed_report)
        print(f"Generated: {failed_file}")

    print()
    print("=== Next Steps ===")
    print(f"1. Process raw data: Run alignment pipeline for {study_id}")
    print(f"2. Create symlinks: bash {script_file}")
    print(f"3. Add JBrowse2 tracks: Merge {tracks_file} into config.json")
    print(f"4. Update expression service: Add config from {expr_file}")
    if failed_samples:
        print(f"\n⚠ Note: {len(failed_samples)} sample(s) excluded due to low alignment quality")


if __name__ == '__main__':
    main()
